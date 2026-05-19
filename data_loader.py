# DATA LOADER
import os, datetime, shutil, tempfile
from openpyxl import load_workbook
from collections import defaultdict

def _get_year(val):
    if val is None: return ""
    if isinstance(val, datetime.datetime): return str(val.year)
    if isinstance(val, (int, float)): return str(int(val))
    return ""

def _fmt_bm(val):
    if val is None: return ""
    if isinstance(val, datetime.datetime): return val.strftime("%b-%y")
    if isinstance(val, (int, float)): return str(int(val))
    return str(val).strip()

def _clean_status(val):
    if val is None: return ""
    if isinstance(val, (int, float)): return "{}%".format(int(round(val * 100)))
    if isinstance(val, datetime.datetime): return val.strftime("%d-%b-%Y")
    return str(val).strip()

def load_dashboard_data(data_folder, tracker_files):
    errors   = []
    targets  = {}
    bb_rows  = []
    pipeline = defaultdict(lambda: defaultdict(float))
    pipe_detail = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    am_order = list(tracker_files.keys())

    for label, fname in tracker_files.items():
        fpath = os.path.join(data_folder, fname)
        if not os.path.exists(fpath):
            errors.append("File not found: " + fpath)
            targets[label] = 0
            continue
        tmp_path = None
        try:
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            tmp.close()
            tmp_path = tmp.name
            shutil.copy2(fpath, tmp_path)
            wb = load_workbook(tmp_path, read_only=True, data_only=True)

            targets[label] = 0
            if "Summary" in wb.sheetnames:
                for row in wb["Summary"].iter_rows(values_only=True):
                    if len(row) > 2 and row[1] == "Top Line Target 2026":
                        v = row[2]
                        targets[label] = float(v) if isinstance(v, (int, float)) else 0
                        break

            if "Booking & Billing Forecast" in wb.sheetnames:
                for i, row in enumerate(wb["Booking & Billing Forecast"].iter_rows(values_only=True)):
                    if i < 2 or len(row) < 15:
                        continue
                    acct, cust, order, bm = row[1], row[3], row[6], row[14]
                    if not acct or not order:
                        continue
                    if isinstance(acct, str) and not acct.strip():
                        continue
                    if isinstance(order, str) and (not order.strip() or order.startswith("=")):
                        continue
                    try:
                        v = float(order)
                    except Exception:
                        continue
                    cust_str = str(cust).strip() if cust else "-"
                    bb_rows.append((label, cust_str, v, _fmt_bm(bm), _get_year(bm)))

            if "Opportunity Tracker" in wb.sheetnames:
                for i, row in enumerate(wb["Opportunity Tracker"].iter_rows(values_only=True)):
                    if i < 1 or len(row) < 12:
                        continue
                    acct, cust, prop, status = row[1], row[3], row[7], row[11]
                    if not acct or not prop:
                        continue
                    if isinstance(acct, str) and not acct.strip():
                        continue
                    if isinstance(prop, str) and (not prop.strip() or prop.startswith("=")):
                        continue
                    try:
                        pv = float(prop)
                    except Exception:
                        continue
                    sc = _clean_status(status)
                    if sc in ("40%", "50%", "80%"):
                        pipeline[label][sc] += pv
                        cust_name = str(cust).strip() if cust else "-"
                        pipe_detail[label][sc][cust_name] += pv

            wb.close()

        except Exception as e:
            errors.append("Error reading {}: {}".format(fname, e))
            targets.setdefault(label, 0)

        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

    actual_2026 = defaultdict(float)
    cust_totals = defaultdict(lambda: defaultdict(float))
    for label, cust, v, bm, yr in bb_rows:
        if yr == "2026":
            actual_2026[label] += v
            cust_totals[label][cust] += v

    accounts_2026 = {}
    for label in am_order:
        consolidated = sorted(cust_totals[label].items(), key=lambda x: -x[1])
        rows = []
        for c, t in consolidated:
            rows.append({"customer": c, "amount": round(t, 2)})
        accounts_2026[label] = rows

    pipeline_detail = {}
    for label in am_order:
        pipeline_detail[label] = {}
        for sc in ("40%", "50%", "80%"):
            items = sorted(pipe_detail[label][sc].items(), key=lambda x: -x[1])
            pipeline_detail[label][sc] = [{"customer": c, "amount": round(v, 2)} for c, v in items]

    total_target      = sum(targets.values())
    total_actual_2026 = sum(actual_2026.values())

    last_updated = None
    for fname in tracker_files.values():
        fpath = os.path.join(data_folder, fname)
        if os.path.exists(fpath):
            mtime = datetime.datetime.fromtimestamp(os.path.getmtime(fpath))
            if last_updated is None or mtime > last_updated:
                last_updated = mtime

    lu_str = last_updated.strftime("%d %b %Y, %H:%M") if last_updated else "Unknown"

    return {
        "total_target":      total_target,
        "total_actual_2026": total_actual_2026,
        "targets":           targets,
        "actual_2026":       dict(actual_2026),
        "accounts_2026":     accounts_2026,
        "pipeline":          {am: dict(pipeline[am]) for am in am_order},
        "pipeline_detail":   pipeline_detail,
        "am_labels":         am_order,
        "last_updated":      lu_str,
        "errors":            errors,
    }
